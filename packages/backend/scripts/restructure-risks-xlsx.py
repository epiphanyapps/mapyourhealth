"""
One-shot script: restructure Risks.xlsx so the 6 miscategorized
contaminants land in their correct sections (so the parser produces the
right category enum without per-row overrides).

Before:
  - Boron, Fluoride, Microcystin-LR, Silver, Aluminium were under
    "27 Chemicals" (parser maps to "organic")
  - Microplastics was under "28 Chemicals - Disinfectants" (parser maps
    to "disinfectant")

After:
  - Silver, Aluminium, Boron, Fluoride → "14 Heavy Metals" (was "10")
    (parser maps "heavy metals" → "inorganic")
  - Microcystin-LR → new "1 Microbiological" section
    (parser CATEGORY_MAP gets a new entry; see parse-risks-excel.ts)
  - Microplastics → "23 Chemicals" (was "27"); other counts adjust

This intentionally loses cell formatting in rows 1–200 of the
"Drinking Water Contamination" sheet — the parser only reads values
and the file is data-driven. Style cleanup is a follow-up if anyone
relies on the formatting.

Run: python3 packages/backend/scripts/restructure-risks-xlsx.py
"""

from __future__ import annotations

import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip3 install openpyxl")

REPO = Path(__file__).resolve().parents[3]
XLSX = REPO / "Risks.xlsx"
BACKUP = REPO / "Risks.before-cascade-fix.xlsx"
SHEET = "Drinking Water Contamination"
WORK_RANGE = 200  # rows we'll restructure; everything else preserved as-is
COLS = 13         # columns A..M


def is_header_row(row: list) -> bool:
    if not isinstance(row[0], str):
        return False
    return bool(re.match(r"^\d+\s+", row[0].strip())) and (len(row) < 2 or not row[1])


def name_of(row: list) -> str | None:
    if isinstance(row[0], str):
        return row[0].strip()
    return None


def update_header_count(header_row: list, new_count: int) -> list:
    """Rewrite '10 Heavy Metals' → '14 Heavy Metals' etc."""
    text = header_row[0]
    new_text = re.sub(r"^\d+", str(new_count), text, count=1)
    return [new_text] + list(header_row[1:])


def main() -> None:
    if not BACKUP.exists():
        sys.exit(f"Backup missing at {BACKUP}; refuse to run without one.")
    if not XLSX.exists():
        sys.exit(f"Source missing: {XLSX}")

    wb = openpyxl.load_workbook(XLSX)
    ws = wb[SHEET]

    # 1. Snapshot rows 1..WORK_RANGE
    snapshot: list[list] = []
    for r in range(1, WORK_RANGE + 1):
        snapshot.append([ws.cell(row=r, column=c).value for c in range(1, COLS + 1)])

    targets = {"Boron", "Fluoride", "Microcystin-LR", "Silver", "Aluminium", "Microplastics"}
    moved: dict[str, list] = {}

    # 2. Walk the snapshot, accumulate sections. Pluck out target rows.
    layout: list[tuple[list | None, list[list]]] = []
    leading_rows: list[list] = []  # rows before the first section header (e.g. "172 contaminants")
    current_header: list | None = None
    current_rows: list[list] = []
    seen_any_header = False

    for row in snapshot:
        if name_of(row) is None:
            continue  # skip blank rows for now; we'll just trim them
        if is_header_row(row):
            if not seen_any_header:
                # Anything before first header (the "172 contaminants" tally row) stays at top
                seen_any_header = True
            else:
                layout.append((current_header, current_rows))
            current_header = row
            current_rows = []
        else:
            if not seen_any_header:
                leading_rows.append(row)
                continue
            n = name_of(row)
            if n in targets:
                moved[n] = row
            else:
                current_rows.append(row)
    if current_header is not None:
        layout.append((current_header, current_rows))

    if set(moved.keys()) != targets:
        missing = targets - set(moved.keys())
        sys.exit(f"Expected to find all 6 targets, missing: {missing}")

    # 3. Inject moves
    new_layout: list[tuple[list, list[list]]] = []
    inserted_microbiological = False
    for header, rows in layout:
        assert header is not None
        title = name_of(header) or ""
        if "Heavy Metals" in title:
            # Append the 4 heavy metals at end of section
            rows = rows + [moved["Silver"], moved["Aluminium"], moved["Boron"], moved["Fluoride"]]
            header = update_header_count(header, len(rows))
            new_layout.append((header, rows))
        elif title.startswith(("27 Chemicals",)) and "Desinfect" not in title and "Disinfect" not in title:
            # Insert NEW microbiological section *before* Chemicals, but only once.
            if not inserted_microbiological:
                micro_header = ["1 Microbiological"] + [None] * (COLS - 1)
                micro_rows = [moved["Microcystin-LR"]]
                micro_header = update_header_count(micro_header, len(micro_rows))
                new_layout.append((micro_header, micro_rows))
                inserted_microbiological = True
            # Add Microplastics to Chemicals
            rows = rows + [moved["Microplastics"]]
            header = update_header_count(header, len(rows))
            new_layout.append((header, rows))
        elif "Desinfect" in title or "Disinfect" in title:
            # Disinfectants: count drops because Microplastics moved out
            header = update_header_count(header, len(rows))
            new_layout.append((header, rows))
        else:
            new_layout.append((header, rows))

    if not inserted_microbiological:
        sys.exit("Could not find '27 Chemicals' section to anchor the new Microbiological section.")

    # 4. Compose final ordered list of rows for the work range
    final_rows: list[list] = []
    final_rows.extend(leading_rows)
    for header, rows in new_layout:
        final_rows.append(header)
        final_rows.extend(rows)

    if len(final_rows) > WORK_RANGE:
        sys.exit(
            f"Restructured row count ({len(final_rows)}) exceeds WORK_RANGE ({WORK_RANGE}); "
            "bump WORK_RANGE and re-verify nothing past row 200 is data."
        )

    # 4b. Unmerge any merged ranges that overlap the work range. Section
    # headers are typically merged across multiple columns ("27 Chemicals"
    # spans A:M); MergedCell.value is read-only, so we have to free them
    # before overwriting.
    overlapping = [
        rng for rng in list(ws.merged_cells.ranges)
        if rng.min_row <= WORK_RANGE
    ]
    for rng in overlapping:
        ws.unmerge_cells(str(rng))
    print(f"Unmerged {len(overlapping)} merged ranges in rows 1..{WORK_RANGE}")

    # 5. Write back: rows 1..WORK_RANGE replaced; everything past is untouched.
    for r in range(1, WORK_RANGE + 1):
        new_values = final_rows[r - 1] if r - 1 < len(final_rows) else [None] * COLS
        for c in range(1, COLS + 1):
            ws.cell(row=r, column=c).value = new_values[c - 1] if c - 1 < len(new_values) else None

    wb.save(XLSX)
    print(f"Wrote {len(final_rows)} rows to {XLSX} (work range 1..{WORK_RANGE}).")
    print("Section summary after restructure:")
    for header, rows in new_layout:
        print(f"  {header[0]!r}: {len(rows)} contaminants")


if __name__ == "__main__":
    main()
